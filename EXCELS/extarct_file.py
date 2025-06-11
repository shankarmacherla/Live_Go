import os
import re
from openpyxl import load_workbook
from PIL import Image
from io import BytesIO

# === CONFIGURATION ===
excel_path = "Ports/ports_2.xlsx"  # Update if needed
output_dir = "Ports/ports_2_images"
os.makedirs(output_dir, exist_ok=True)

# === Helper to sanitize filenames ===
def sanitize_filename(name):
    name = name.replace(" ", "_")
    name = re.sub(r'[\\/*?:"<>|\n\r\t]', '', name)
    return name.strip()

# === LOAD WORKBOOK ===
wb = load_workbook(excel_path)
ws = wb.active

# === MAP IMAGES TO ROWS ===
img_map = {}
for img in ws._images:
    row = img.anchor._from.row + 1
    img_map[row] = img

# === ITERATE ROWS AND SAVE IMAGES ===
for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    name = row[0].value  # Assuming Name is in column A
    if not name:
        continue

    name_sanitized = sanitize_filename(name)
    if idx in img_map:
        img = img_map[idx]
        img_data = img._data()
        image = Image.open(BytesIO(img_data))

        # 🛠️ Convert 'P' or 'RGBA' to 'RGB'
        if image.mode in ('RGBA', 'P'):
            image = image.convert('RGB')

        out_path = os.path.join(output_dir, f"{name_sanitized}.jpg")
        image.save(out_path)
        print(f"✅ Saved: {os.path.basename(out_path)}")
    else:
        print(f"⚠️ No image found for: {name}")
