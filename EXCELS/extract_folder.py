import os
import re
from openpyxl import load_workbook
from PIL import Image
from io import BytesIO

# === CONFIGURATION ===
excel_folder = "Rack/Rack Vendors"   # 🔁 Your folder path
output_base_dir = "Rack/RackVendors_Images"
os.makedirs(output_base_dir, exist_ok=True)

# === Helper to sanitize filenames ===
def sanitize_filename(name):
    name = name.replace(" ", "_")
    name = re.sub(r'[\\/*?:"<>|\n\r\t]', '', name)
    return name.strip()

# === Process Each Excel File ===
for filename in os.listdir(excel_folder):
    if not filename.endswith(".xlsx"):
        continue

    excel_path = os.path.join(excel_folder, filename)
    print(f"📄 Processing: {filename}")

    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        # Build a map of row to image
        img_map = {}
        for img in ws._images:
            row = img.anchor._from.row + 1
            img_map[row] = img

        # Subfolder for this Excel's images
        excel_output_dir = os.path.join(output_base_dir, os.path.splitext(filename)[0])
        os.makedirs(excel_output_dir, exist_ok=True)

        # Iterate over rows and save images
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            name = row[0].value  # 'Name' column
            if not name:
                continue

            name_sanitized = sanitize_filename(name)
            if idx in img_map:
                img = img_map[idx]
                img_data = img._data()
                image = Image.open(BytesIO(img_data))

                # Convert RGBA or P mode to RGB
                if image.mode in ('RGBA', 'P'):
                    image = image.convert('RGB')

                out_path = os.path.join(excel_output_dir, f"{name_sanitized}.jpg")
                image.save(out_path)
                print(f"✅ Saved: {out_path}")
            else:
                print(f"⚠️ No image found for row {idx} in {filename}")
    except Exception as e:
        print(f"❌ Failed to process {filename}: {e}")
